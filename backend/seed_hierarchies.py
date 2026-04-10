"""
Seed script: reads Frameworks_Hierarchy.xlsx and imports all nodes into the database.
Run from within the backend container or with the right DATABASE_URL.
"""
import asyncio
import uuid
import sys
import os

# Add project root to path
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from decimal import Decimal
from sqlalchemy import select, delete

from app.database import async_session, engine, Base
from app.models.compliance_framework import ComplianceFramework
from app.models.regulatory_entity import RegulatoryEntity
from app.models.framework_node import FrameworkNode, NodeType

EXCEL_PATH = "/app/Frameworks_Hierarchy.xlsx"

# Map sheet names to framework abbreviations
SHEET_TO_FRAMEWORK = {
    "National AI Index (NAII)": "NAII",
    "National Data Index (NDI)": "NDI",
    "IT Governance Framework (ITGF)": "ITGF",
}

# Node types to seed for ITGF (new framework)
ITGF_NODE_TYPES = [
    ("domain", "Domain", "#00338D", False),
    ("subdomain", "Sub-Domain", "#0091DA", False),
    ("control", "Control", "#27AE60", True),
]


async def ensure_itgf_framework(db):
    """Create ITGF framework if it doesn't exist."""
    result = await db.execute(select(ComplianceFramework).where(ComplianceFramework.abbreviation == "ITGF"))
    fw = result.scalar_one_or_none()
    if fw:
        return fw

    # Get DGA entity (or SDAIA as fallback)
    entity = (await db.execute(select(RegulatoryEntity).where(RegulatoryEntity.abbreviation == "DGA"))).scalar_one_or_none()
    if not entity:
        entity = (await db.execute(select(RegulatoryEntity))).scalars().first()

    fw = ComplianceFramework(
        name="IT Governance Framework",
        abbreviation="ITGF",
        name_ar="إطار حوكمة تقنية المعلومات",
        description="IT Governance Framework for Saudi government entities.",
        entity_id=entity.id,
        status="Active",
        icon="monitor",
    )
    db.add(fw)
    await db.flush()

    # Add node types
    for i, (name, label, color, assessable) in enumerate(ITGF_NODE_TYPES):
        db.add(NodeType(framework_id=fw.id, name=name, label=label, color=color, sort_order=i, is_assessable_default=assessable))
    await db.flush()

    print(f"  Created ITGF framework: {fw.id}")
    return fw


async def import_sheet(db, sheet_name, ws):
    fw_abbr = SHEET_TO_FRAMEWORK.get(sheet_name)
    if not fw_abbr:
        print(f"  SKIP: Unknown sheet '{sheet_name}'")
        return

    # Get framework
    result = await db.execute(select(ComplianceFramework).where(ComplianceFramework.abbreviation == fw_abbr))
    framework = result.scalar_one_or_none()
    if not framework:
        print(f"  ERROR: Framework '{fw_abbr}' not found")
        return

    # Clear existing nodes for this framework
    await db.execute(delete(FrameworkNode).where(FrameworkNode.framework_id == framework.id))
    await db.flush()

    # Parse rows: Code | Name (EN) | Name (AR) | Node Type | Depth | Description (EN) | Description (AR) | Guidance (EN) | Guidance (AR)
    ref_to_id = {}
    ref_to_node = {}
    count = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        code = str(row[0]).strip() if row[0] else None
        if not code:
            continue

        name_en = str(row[1]).strip() if row[1] else code
        name_ar = str(row[2]).strip() if row[2] else None
        node_type = str(row[3]).strip() if row[3] else "domain"
        depth = int(row[4]) if row[4] is not None else 0
        desc_en = str(row[5]).strip() if row[5] else None
        desc_ar = str(row[6]).strip() if row[6] else None
        guidance_en = str(row[7]).strip() if row[7] else None
        guidance_ar = str(row[8]).strip() if row[8] else None

        # Determine parent from code structure
        parent_id = None
        if depth > 0:
            # Find parent by walking up the code hierarchy
            # e.g., P1.1.1 -> parent is P1.1, DG.MQ.1 -> parent is DG
            parts = code.rsplit(".", 1)
            if len(parts) > 1:
                parent_code = parts[0]
                # Try exact match
                if parent_code in ref_to_id:
                    parent_id = ref_to_id[parent_code]
                else:
                    # Try removing last segment more aggressively
                    # For codes like MCM.MQ.1, parent might be MCM
                    segments = code.split(".")
                    for i in range(len(segments) - 1, 0, -1):
                        candidate = ".".join(segments[:i])
                        if candidate in ref_to_id:
                            parent_id = ref_to_id[candidate]
                            break

            if parent_id is None and depth > 0:
                # Fallback: find last node at depth-1
                for prev_code, prev_node in reversed(list(ref_to_node.items())):
                    if prev_node["depth"] == depth - 1:
                        parent_id = ref_to_id[prev_code]
                        break

        node_id = uuid.uuid4()

        # Calculate path
        if parent_id and parent_id in [ref_to_id[c] for c in ref_to_id]:
            parent_path = None
            for c, nid in ref_to_id.items():
                if nid == parent_id:
                    parent_path = ref_to_node[c]["path"]
                    break
            path = f"{parent_path}{node_id}/" if parent_path else f"/{node_id}/"
        else:
            path = f"/{node_id}/"

        # Determine if assessable (leaf nodes at max depth)
        is_assessable = node_type in ("question", "control", "indicator", "specification")

        node = FrameworkNode(
            id=node_id,
            framework_id=framework.id,
            parent_id=parent_id,
            node_type=node_type,
            reference_code=code,
            name=name_en.lstrip(),  # Remove leading spaces used for visual indentation
            name_ar=name_ar,
            description=desc_en,
            description_ar=desc_ar,
            guidance=guidance_en,
            guidance_ar=guidance_ar,
            sort_order=count,
            path=path,
            depth=depth,
            is_active=True,
            is_assessable=is_assessable,
        )
        db.add(node)

        ref_to_id[code] = node_id
        ref_to_node[code] = {"depth": depth, "path": path}
        count += 1

    await db.flush()
    print(f"  {fw_abbr}: imported {count} nodes")


async def main():
    print("=== Seeding Framework Hierarchies ===")
    print(f"Reading: {EXCEL_PATH}")

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    print(f"Sheets: {wb.sheetnames}")

    async with async_session() as db:
        # Ensure ITGF framework exists
        await ensure_itgf_framework(db)
        await db.commit()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        async with async_session() as db:
            await import_sheet(db, sheet_name, ws)
            await db.commit()

    print("\n=== Done ===")


if __name__ == "__main__":
    asyncio.run(main())
