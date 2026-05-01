"""AI Assessment API — LLM Models CRUD + AI Assess endpoints."""
import json
import time
import uuid
from datetime import datetime, timezone
from decimal import Decimal

import httpx
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from pydantic import BaseModel
from sqlalchemy import select, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.core.permissions import require_role
from app.database import get_db
from app.dependencies import get_current_user
from app.models.llm_model import LlmModel
from app.models.ai_assessment_log import AiAssessmentLog
from app.models.assessment_engine import AssessmentInstance, AssessmentResponse, AssessmentEvidence
from app.models.user import User
from app.services import ollama_bootstrap

router = APIRouter(tags=["ai-assessment"])


# ============ LLM MODELS CRUD ============

class LlmModelCreate(BaseModel):
    name: str
    model_id: str
    max_tokens: int = 4096
    temperature: float = 0.0
    context_window: int = 8192
    supports_documents: bool = False
    is_default: bool = False
    description: str | None = None

class BulkDeleteModelsRequest(BaseModel):
    ids: list[uuid.UUID]

def _model_resp(m: LlmModel) -> dict:
    return {
        "id": str(m.id), "name": m.name, "provider": "ollama", "model_id": m.model_id,
        "max_tokens": m.max_tokens, "temperature": float(m.temperature),
        "context_window": m.context_window, "supports_documents": m.supports_documents,
        "is_default": m.is_default, "is_active": m.is_active, "description": m.description,
        "last_tested_at": m.last_tested_at.isoformat() if m.last_tested_at else None,
        "created_at": m.created_at.isoformat() if m.created_at else "",
    }

@router.get("/api/settings/llm-models")
async def list_models(db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(LlmModel).where(LlmModel.is_active == True).order_by(LlmModel.name))
    return [_model_resp(m) for m in result.scalars().all()]

# Static routes MUST come before {model_id} parameterized routes
@router.get("/api/settings/llm-models/export-excel")
async def export_models_excel(db: AsyncSession = Depends(get_db), current_user: User = Depends(require_role("admin"))):
    from io import BytesIO
    from openpyxl import Workbook
    from fastapi.responses import StreamingResponse
    models = (await db.execute(select(LlmModel).where(LlmModel.is_active == True).order_by(LlmModel.name))).scalars().all()
    wb = Workbook()
    ws = wb.active
    ws.title = "LLM Models"
    headers = ["name", "model_id", "max_tokens", "temperature", "context_window", "supports_documents", "is_default", "description"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    for row_idx, m in enumerate(models, 2):
        vals = [m.name, m.model_id, m.max_tokens, float(m.temperature), m.context_window, m.supports_documents, m.is_default, m.description]
        for i, v in enumerate(vals, 1):
            ws.cell(row=row_idx, column=i, value=v)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=llm_models.xlsx"})

@router.post("/api/settings/llm-models/import-excel")
async def import_models_excel(file: UploadFile = File(...), preview: bool = False, db: AsyncSession = Depends(get_db), current_user: User = Depends(require_role("admin"))):
    from io import BytesIO
    from openpyxl import load_workbook
    try:
        content = await file.read()
        if not content:
            raise HTTPException(400, "Empty file uploaded")
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        header_row = [cell.value for cell in ws[1]]
        if not header_row or "name" not in header_row:
            raise HTTPException(400, "Invalid file format: missing 'name' column header")
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            r = dict(zip(header_row, row))
            if r.get("name"):
                rows.append(r)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Failed to parse Excel file: {str(e)}")
    existing_names = set((await db.execute(select(LlmModel.name).where(LlmModel.is_active == True))).scalars().all())
    new_rows = [r for r in rows if r["name"] not in existing_names]
    dup_rows = [r for r in rows if r["name"] in existing_names]
    if preview:
        return {
            "total_in_file": len(rows),
            "new_items": [{"name": r["name"], "model_id": r.get("model_id", "")} for r in new_rows],
            "duplicates": [{"name": r["name"]} for r in dup_rows],
            "will_import": len(new_rows),
            "will_skip": len(dup_rows),
        }
    created = 0
    for r in new_rows:
        m = LlmModel(
            name=r["name"],
            model_id=r.get("model_id") or "",
            max_tokens=int(r.get("max_tokens") or 4096),
            temperature=Decimal(str(r.get("temperature") or 0.0)),
            context_window=int(r.get("context_window") or 8192),
            supports_documents=bool(r.get("supports_documents")),
            is_default=bool(r.get("is_default")),
            description=r.get("description"),
        )
        db.add(m)
        created += 1
    await db.flush()
    return {"imported": created, "skipped_duplicates": len(dup_rows)}

@router.post("/api/settings/llm-models/bulk-delete")
async def bulk_delete_models(data: BulkDeleteModelsRequest, db: AsyncSession = Depends(get_db), current_user: User = Depends(require_role("admin"))):
    """Soft-delete multiple LLM models at once."""
    if not data.ids:
        raise HTTPException(400, "No model IDs provided")
    result = await db.execute(select(LlmModel).where(LlmModel.id.in_(data.ids)))
    models = result.scalars().all()
    if not models:
        raise HTTPException(404, "No matching models found")
    deleted = 0
    already_removed = 0
    had_default = False
    for m in models:
        if m.is_default:
            had_default = True
        if m.is_active:
            m.is_active = False
            deleted += 1
        else:
            already_removed += 1
    await db.flush()
    return {
        "deleted": deleted,
        "already_removed": already_removed,
        "requested": len(data.ids),
        "had_default": had_default,
    }

@router.get("/api/settings/llm-models/{model_id}")
async def get_model(model_id: uuid.UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    m = (await db.execute(select(LlmModel).where(LlmModel.id == model_id))).scalar_one_or_none()
    if not m: raise HTTPException(404, "Model not found")
    return _model_resp(m)

@router.post("/api/settings/llm-models", status_code=201)
async def create_model(data: LlmModelCreate, db: AsyncSession = Depends(get_db), current_user: User = Depends(require_role("admin"))):
    if data.is_default:
        await db.execute(update(LlmModel).values(is_default=False))
    m = LlmModel(
        name=data.name, model_id=data.model_id,
        max_tokens=data.max_tokens, temperature=Decimal(str(data.temperature)),
        context_window=data.context_window, supports_documents=data.supports_documents,
        is_default=data.is_default, description=data.description,
    )
    db.add(m); await db.flush(); await db.refresh(m)
    return _model_resp(m)

@router.put("/api/settings/llm-models/{model_id}")
async def update_model(model_id: uuid.UUID, data: LlmModelCreate, db: AsyncSession = Depends(get_db), current_user: User = Depends(require_role("admin"))):
    m = (await db.execute(select(LlmModel).where(LlmModel.id == model_id))).scalar_one_or_none()
    if not m: raise HTTPException(404, "Model not found")
    if data.is_default and not m.is_default:
        await db.execute(update(LlmModel).where(LlmModel.id != model_id).values(is_default=False))
    for k, v in data.model_dump(exclude_unset=True).items():
        if k == "temperature": v = Decimal(str(v))
        setattr(m, k, v)
    await db.flush(); await db.refresh(m)
    return _model_resp(m)

@router.delete("/api/settings/llm-models/{model_id}", status_code=204)
async def delete_model(model_id: uuid.UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(require_role("admin"))):
    m = (await db.execute(select(LlmModel).where(LlmModel.id == model_id))).scalar_one_or_none()
    if not m: raise HTTPException(404, "Model not found")
    m.is_active = False; await db.flush()

@router.post("/api/settings/llm-models/{model_id}/test")
async def test_model(model_id: uuid.UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(require_role("admin"))):
    m = (await db.execute(select(LlmModel).where(LlmModel.id == model_id))).scalar_one_or_none()
    if not m: raise HTTPException(404, "Model not found")
    start = time.time()
    try:
        response_text = await _call_llm(m, "You are a test assistant.", "Say 'Hello, I am working correctly.' in exactly those words.", force_deterministic=True)
        elapsed = int((time.time() - start) * 1000)
        m.last_tested_at = datetime.now(timezone.utc)
        await db.flush()
        return {"success": True, "response": response_text[:200], "response_time_ms": elapsed}
    except Exception as e:
        return {"success": False, "error": str(e), "response_time_ms": int((time.time() - start) * 1000)}


# ============ OLLAMA DISCOVERY + AUTO-SETUP ============

@router.get("/api/settings/ollama/models")
async def list_installed_ollama_models(current_user: User = Depends(get_current_user)):
    """Return models currently pulled into the host's Ollama instance."""
    models = await ollama_bootstrap.list_ollama_models()
    return [
        {
            "name": m.get("name"),
            "size_gb": round((m.get("size") or 0) / (1024 ** 3), 2),
            "modified_at": m.get("modified_at"),
        }
        for m in models if m.get("name")
    ]

@router.get("/api/settings/ollama/status")
async def ollama_status(current_user: User = Depends(get_current_user)):
    """Reachability + default-model health, for the UI status banner."""
    return {
        **ollama_bootstrap.get_status(),
        "base_url": settings.OLLAMA_BASE_URL,
    }

@router.post("/api/settings/llm-models/auto-setup")
async def trigger_auto_setup(db: AsyncSession = Depends(get_db), current_user: User = Depends(require_role("admin"))):
    """Manually trigger the Ollama auto-setup flow (same logic as startup)."""
    return await ollama_bootstrap.ensure_default(db)


# ============ AI ASSESSMENT ENDPOINTS ============

class AssessRequest(BaseModel):
    model_id: uuid.UUID | None = None
    ai_product_id: uuid.UUID | None = None

@router.post("/api/assessments/{inst_id}/ai-assess/{node_id}")
async def assess_node(inst_id: uuid.UUID, node_id: uuid.UUID, data: AssessRequest = AssessRequest(), db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    from app.services.prompt_assembly import build_prompt

    # Get model
    if data.model_id:
        model = (await db.execute(select(LlmModel).where(LlmModel.id == data.model_id, LlmModel.is_active == True))).scalar_one_or_none()
    else:
        model = (await db.execute(select(LlmModel).where(LlmModel.is_default == True, LlmModel.is_active == True))).scalar_one_or_none()
    if not model:
        raise HTTPException(400, "No AI model configured. Go to Settings → LLM Models to register one.")

    # Build prompts dynamically from database
    prompts = await build_prompt(db, inst_id, node_id, ai_product_id=data.ai_product_id)
    system_prompt = prompts["system_prompt"]
    user_prompt = prompts["user_prompt"]

    # Call LLM
    start = time.time()
    error_msg = None
    raw_response = None
    parsed = None
    try:
        # Document assessment MUST run at temperature=0 for reproducibility.
        raw_response = await _call_llm(model, system_prompt, user_prompt, force_deterministic=True)
        parsed = _parse_suggestion(raw_response)
    except Exception as e:
        error_msg = str(e)

    elapsed = int((time.time() - start) * 1000)

    # Log
    log = AiAssessmentLog(
        instance_id=inst_id, node_id=node_id, model_id=model.id,
        system_prompt_sent=system_prompt, user_prompt_sent=user_prompt,
        raw_response=raw_response, parsed_suggestion=parsed,
        processing_time_ms=elapsed, error=error_msg,
    )
    db.add(log)
    await db.flush()

    if error_msg:
        raise HTTPException(500, f"AI assessment failed: {error_msg}")

    return {
        "suggestion": parsed,
        "model_name": model.name,
        "processing_time_ms": elapsed,
        "log_id": str(log.id),
    }


class AcceptRequest(BaseModel):
    suggestion: dict

@router.post("/api/assessments/{inst_id}/ai-assess/{node_id}/accept")
async def accept_suggestion(inst_id: uuid.UUID, node_id: uuid.UUID, data: AcceptRequest, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    ai_product_id = data.suggestion.get("_ai_product_id")
    q = select(AssessmentResponse).where(AssessmentResponse.instance_id == inst_id, AssessmentResponse.node_id == node_id)
    if ai_product_id:
        q = q.where(AssessmentResponse.ai_product_id == uuid.UUID(ai_product_id))
    else:
        q = q.where(AssessmentResponse.ai_product_id.is_(None))
    resp = (await db.execute(q)).scalar_one_or_none()
    if not resp: raise HTTPException(404, "Response not found")

    suggestion = data.suggestion
    rd = dict(resp.response_data) if resp.response_data else {}
    rd["review_feedback"] = suggestion.get("review_feedback", "")
    rd["justification"] = suggestion.get("justification", "")
    rd["recommendation"] = suggestion.get("recommendations", "")
    rd["compliance_status"] = suggestion.get("compliance_status", "")
    rd["ai_assessment"] = suggestion
    # Store document verification as human-readable status strings
    doc_analysis = suggestion.get("document_analysis", [])
    if doc_analysis:
        all_approved = all(d.get("is_approved", False) for d in doc_analysis)
        rd["doc_approved"] = "Approved" if all_approved else "Not Approved"
        all_signed = all(d.get("has_signature", False) for d in doc_analysis)
        rd["doc_signed"] = "Signed" if all_signed else "No Signature"
        has_date = any(d.get("document_date") for d in doc_analysis)
        if has_date:
            dates = [d.get("document_date") for d in doc_analysis if d.get("document_date")]
            rd["doc_date_check"] = "Valid Date"
            # Store the actual date for reference
            rd["doc_date_value"] = dates[0] if dates else None
        else:
            rd["doc_date_check"] = "No Date"
    else:
        rd["doc_approved"] = "Not Checked"
        rd["doc_signed"] = "Not Checked"
        rd["doc_date_check"] = "Not Checked"
    resp.response_data = rd
    resp.status = "draft"
    resp.computed_score_label = suggestion.get("compliance_status")
    resp.scored_by = current_user.id
    resp.scored_at = datetime.now(timezone.utc)

    # Mark log as accepted
    await db.execute(update(AiAssessmentLog).where(
        AiAssessmentLog.instance_id == inst_id, AiAssessmentLog.node_id == node_id, AiAssessmentLog.accepted.is_(None)
    ).values(accepted=True))

    # Update evidence metadata from AI document_analysis
    doc_analysis = suggestion.get("document_analysis", [])
    if doc_analysis and resp:
        from app.models.assessment_engine import AssessmentEvidence
        evs = (await db.execute(select(AssessmentEvidence).where(AssessmentEvidence.response_id == resp.id))).scalars().all()
        ev_map = {e.file_name: e for e in evs}
        for da in doc_analysis:
            ev = ev_map.get(da.get("file_name"))
            if not ev:
                continue
            if da.get("document_date"):
                try:
                    from datetime import date as date_type
                    ev.document_date = date_type.fromisoformat(da["document_date"])
                except (ValueError, TypeError):
                    pass
            if da.get("is_approved") is not None:
                ev.is_approved = da["is_approved"]
            if da.get("approved_by"):
                ev.approved_by = da["approved_by"]
            if da.get("has_signature") is not None:
                ev.has_signature = da["has_signature"]
            if da.get("notes"):
                ev.reviewer_notes = da["notes"]

    await db.flush()
    return {"status": "accepted", "compliance_status": suggestion.get("compliance_status")}


@router.post("/api/assessments/{inst_id}/ai-assess/{node_id}/dismiss")
async def dismiss_suggestion(inst_id: uuid.UUID, node_id: uuid.UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    await db.execute(update(AiAssessmentLog).where(
        AiAssessmentLog.instance_id == inst_id, AiAssessmentLog.node_id == node_id, AiAssessmentLog.accepted.is_(None)
    ).values(accepted=False))
    await db.flush()
    return {"status": "dismissed"}


# ============ LLM CALL HELPER ============

async def _call_llm(
    model: LlmModel,
    system_prompt: str,
    user_prompt: str,
    *,
    force_deterministic: bool,
) -> str:
    """Call Ollama. `force_deterministic=True` pins temperature=0 regardless of the DB row.

    Always pass `force_deterministic=True` for document-assessment paths — that
    is the only way results are reproducible against acceptance criteria.
    """
    temperature = 0.0 if force_deterministic else float(model.temperature or 0)
    endpoint = f"{settings.OLLAMA_BASE_URL.rstrip('/')}/api/chat"
    payload = {
        "model": model.model_id,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "stream": False,
        "options": {"temperature": temperature},
    }
    async with httpx.AsyncClient(timeout=120) as client:
        resp = await client.post(endpoint, json=payload, headers={"Content-Type": "application/json"})
        resp.raise_for_status()
        return resp.json()["message"]["content"]


def _parse_suggestion(raw: str) -> dict:
    """Parse LLM response — handle markdown code fences."""
    text = raw.strip()
    if text.startswith("```"):
        text = text.split("\n", 1)[1] if "\n" in text else text[3:]
        if text.endswith("```"):
            text = text[:-3]
        text = text.strip()
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        start = text.find("{")
        end = text.rfind("}") + 1
        if start >= 0 and end > start:
            parsed = json.loads(text[start:end])
        else:
            raise ValueError(f"Could not parse JSON from LLM response: {text[:200]}")

    required = ["compliance_status", "review_feedback", "justification"]
    for field in required:
        if field not in parsed:
            raise ValueError(f"Missing required field in suggestion: {field}")
    # Normalize compliance_status
    status = parsed.get("compliance_status", "").strip()
    valid = {"Compliant", "Semi-Compliant", "Non-Compliant"}
    if status not in valid:
        # Try fuzzy match
        lower = status.lower()
        if "non" in lower:
            parsed["compliance_status"] = "Non-Compliant"
        elif "semi" in lower or "partial" in lower:
            parsed["compliance_status"] = "Semi-Compliant"
        else:
            parsed["compliance_status"] = "Compliant"
    return parsed
