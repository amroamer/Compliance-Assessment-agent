"""Regulator Feedback API — CRUD, bulk upload/export, corrections, evidence, and stats."""
import os
import uuid
from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO
from typing import Optional

import aiofiles
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.permissions import require_role
from app.database import get_db
from app.dependencies import get_current_user
from app.models.assessment_engine import (
    AssessmentInstance, AssessmentResponse, AssessmentScaleLevel,
)
from app.models.cycle_phase import AssessmentCyclePhase
from app.models.framework_node import FrameworkNode
from app.models.regulator_feedback import RegulatorEvidence, RegulatorFeedback
from app.models.user import User

router = APIRouter(tags=["regulator-feedback"])

REGULATOR_EVIDENCE_DIR = os.environ.get("REGULATOR_EVIDENCE_DIR", "/app/uploads/regulator")

# ============ SCHEMAS ============

class FeedbackCreate(BaseModel):
    node_id: uuid.UUID
    phase_id: uuid.UUID
    agreement_status: str = "not_reviewed"
    regulator_score: float | None = None
    regulator_score_label: str | None = None
    feedback_text: str | None = None
    required_actions: list = []
    priority: str = "observation"

class FeedbackUpdate(BaseModel):
    agreement_status: str | None = None
    regulator_score: float | None = None
    regulator_score_label: str | None = None
    feedback_text: str | None = None
    required_actions: list | None = None
    priority: str | None = None

class CorrectionUpdate(BaseModel):
    correction_status: str
    correction_notes: str | None = None

class ActionToggle(BaseModel):
    addressed: bool


# ============ HELPERS ============

def _feedback_resp(fb: RegulatorFeedback, entity_score=None, entity_score_label=None):
    score_gap = None
    if fb.regulator_score is not None and entity_score is not None:
        score_gap = float(entity_score) - float(fb.regulator_score)
    return {
        "id": str(fb.id),
        "instance_id": str(fb.instance_id),
        "node": {
            "id": str(fb.node.id),
            "reference_code": fb.node.reference_code,
            "name": fb.node.name,
            "name_ar": fb.node.name_ar,
            "node_type": fb.node.node_type,
        } if fb.node else None,
        "phase": {
            "id": str(fb.phase.id),
            "name": fb.phase.name,
            "phase_number": fb.phase.phase_number,
        } if fb.phase else None,
        "entity_score": float(entity_score) if entity_score is not None else None,
        "entity_score_label": entity_score_label,
        "regulator_score": float(fb.regulator_score) if fb.regulator_score is not None else None,
        "regulator_score_label": fb.regulator_score_label,
        "score_gap": round(score_gap, 2) if score_gap is not None else None,
        "agreement_status": fb.agreement_status,
        "feedback_text": fb.feedback_text,
        "required_actions": fb.required_actions or [],
        "priority": fb.priority,
        "feedback_by": str(fb.feedback_by) if fb.feedback_by else None,
        "feedback_at": fb.feedback_at.isoformat() if fb.feedback_at else None,
        "updated_at": fb.updated_at.isoformat() if fb.updated_at else None,
        "correction_status": fb.correction_status,
        "correction_notes": fb.correction_notes,
        "corrected_at": fb.corrected_at.isoformat() if fb.corrected_at else None,
        "evidence_files": [
            {"id": str(e.id), "file_name": e.file_name, "file_size": e.file_size,
             "download_url": f"/api/assessments/regulator-evidence/{e.id}/download"}
            for e in (fb.evidence_files or [])
        ],
    }


async def _get_entity_score(db: AsyncSession, instance_id: uuid.UUID, node_id: uuid.UUID):
    """Get the entity's self-assessment score for a node."""
    resp = (await db.execute(
        select(AssessmentResponse)
        .where(AssessmentResponse.instance_id == instance_id, AssessmentResponse.node_id == node_id)
        .limit(1)
    )).scalar_one_or_none()
    if resp:
        return resp.computed_score, resp.computed_score_label
    return None, None


# ============ FEEDBACK CRUD ============

@router.get("/api/assessments/{inst_id}/regulator-feedback")
async def list_feedback(
    inst_id: uuid.UUID,
    phase_id: uuid.UUID | None = None,
    agreement_status: str | None = None,
    priority: str | None = None,
    correction_status: str | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = select(RegulatorFeedback).where(RegulatorFeedback.instance_id == inst_id)
    if phase_id:
        query = query.where(RegulatorFeedback.phase_id == phase_id)
    if agreement_status:
        query = query.where(RegulatorFeedback.agreement_status == agreement_status)
    if priority:
        query = query.where(RegulatorFeedback.priority == priority)
    if correction_status:
        query = query.where(RegulatorFeedback.correction_status == correction_status)
    feedbacks = (await db.execute(query.order_by(RegulatorFeedback.feedback_at))).scalars().all()
    results = []
    for fb in feedbacks:
        es, esl = await _get_entity_score(db, inst_id, fb.node_id)
        results.append(_feedback_resp(fb, es, esl))
    return results


@router.get("/api/assessments/{inst_id}/regulator-feedback/by-node/{node_id}")
async def get_feedback_for_node(
    inst_id: uuid.UUID, node_id: uuid.UUID,
    db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user),
):
    feedbacks = (await db.execute(
        select(RegulatorFeedback)
        .where(RegulatorFeedback.instance_id == inst_id, RegulatorFeedback.node_id == node_id)
        .order_by(RegulatorFeedback.feedback_at)
    )).scalars().all()
    es, esl = await _get_entity_score(db, inst_id, node_id)
    return [_feedback_resp(fb, es, esl) for fb in feedbacks]


@router.post("/api/assessments/{inst_id}/regulator-feedback", status_code=201)
async def create_or_update_feedback(
    inst_id: uuid.UUID, data: FeedbackCreate,
    db: AsyncSession = Depends(get_db), current_user: User = Depends(require_role("admin")),
):
    inst = (await db.execute(select(AssessmentInstance).where(AssessmentInstance.id == inst_id))).scalar_one_or_none()
    if not inst:
        raise HTTPException(404, "Assessment instance not found")

    # Upsert: check if feedback exists for this instance+node+phase
    existing = (await db.execute(
        select(RegulatorFeedback).where(
            RegulatorFeedback.instance_id == inst_id,
            RegulatorFeedback.node_id == data.node_id,
            RegulatorFeedback.phase_id == data.phase_id,
        )
    )).scalar_one_or_none()

    # Auto-resolve score label from scale if not provided
    score_label = data.regulator_score_label
    if data.regulator_score is not None and not score_label:
        score_label = await _resolve_score_label(db, inst.framework_id, data.regulator_score)

    if existing:
        existing.agreement_status = data.agreement_status
        existing.regulator_score = Decimal(str(data.regulator_score)) if data.regulator_score is not None else None
        existing.regulator_score_label = score_label
        existing.feedback_text = data.feedback_text
        existing.required_actions = data.required_actions
        existing.priority = data.priority
        existing.updated_by = current_user.id
        await db.flush()
        es, esl = await _get_entity_score(db, inst_id, data.node_id)
        return _feedback_resp(existing, es, esl)

    fb = RegulatorFeedback(
        instance_id=inst_id,
        node_id=data.node_id,
        phase_id=data.phase_id,
        agreement_status=data.agreement_status,
        regulator_score=Decimal(str(data.regulator_score)) if data.regulator_score is not None else None,
        regulator_score_label=score_label,
        feedback_text=data.feedback_text,
        required_actions=data.required_actions,
        priority=data.priority,
        feedback_by=current_user.id,
    )
    db.add(fb)
    await db.flush()
    es, esl = await _get_entity_score(db, inst_id, data.node_id)
    return _feedback_resp(fb, es, esl)


async def _resolve_score_label(db: AsyncSession, framework_id: uuid.UUID, score_value: float) -> str | None:
    """Try to find a matching scale level label for a score value."""
    level = (await db.execute(
        select(AssessmentScaleLevel)
        .join(AssessmentScaleLevel.scale)
        .where(AssessmentScaleLevel.value == Decimal(str(score_value)))
        .limit(1)
    )).scalar_one_or_none()
    return level.label if level else None


@router.put("/api/assessments/{inst_id}/regulator-feedback/{feedback_id}")
async def update_feedback(
    inst_id: uuid.UUID, feedback_id: uuid.UUID, data: FeedbackUpdate,
    db: AsyncSession = Depends(get_db), current_user: User = Depends(require_role("admin")),
):
    fb = (await db.execute(
        select(RegulatorFeedback).where(RegulatorFeedback.id == feedback_id, RegulatorFeedback.instance_id == inst_id)
    )).scalar_one_or_none()
    if not fb:
        raise HTTPException(404, "Feedback not found")
    updates = data.model_dump(exclude_none=True)
    if "regulator_score" in updates and updates["regulator_score"] is not None:
        updates["regulator_score"] = Decimal(str(updates["regulator_score"]))
    for k, v in updates.items():
        setattr(fb, k, v)
    fb.updated_by = current_user.id
    await db.flush()
    es, esl = await _get_entity_score(db, inst_id, fb.node_id)
    return _feedback_resp(fb, es, esl)


@router.delete("/api/assessments/{inst_id}/regulator-feedback/{feedback_id}", status_code=204)
async def delete_feedback(
    inst_id: uuid.UUID, feedback_id: uuid.UUID,
    db: AsyncSession = Depends(get_db), current_user: User = Depends(require_role("admin")),
):
    fb = (await db.execute(
        select(RegulatorFeedback).where(RegulatorFeedback.id == feedback_id, RegulatorFeedback.instance_id == inst_id)
    )).scalar_one_or_none()
    if not fb:
        raise HTTPException(404, "Feedback not found")
    await db.delete(fb)
    await db.flush()


# ============ CORRECTION TRACKING ============

@router.put("/api/assessments/{inst_id}/regulator-feedback/{feedback_id}/correction")
async def update_correction(
    inst_id: uuid.UUID, feedback_id: uuid.UUID, data: CorrectionUpdate,
    db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user),
):
    fb = (await db.execute(
        select(RegulatorFeedback).where(RegulatorFeedback.id == feedback_id, RegulatorFeedback.instance_id == inst_id)
    )).scalar_one_or_none()
    if not fb:
        raise HTTPException(404, "Feedback not found")
    fb.correction_status = data.correction_status
    fb.correction_notes = data.correction_notes
    fb.corrected_at = datetime.now(timezone.utc)
    fb.corrected_by = current_user.id
    await db.flush()
    es, esl = await _get_entity_score(db, inst_id, fb.node_id)
    return _feedback_resp(fb, es, esl)


@router.put("/api/assessments/{inst_id}/regulator-feedback/{feedback_id}/actions/{action_idx}/toggle")
async def toggle_action(
    inst_id: uuid.UUID, feedback_id: uuid.UUID, action_idx: int, data: ActionToggle,
    db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user),
):
    fb = (await db.execute(
        select(RegulatorFeedback).where(RegulatorFeedback.id == feedback_id, RegulatorFeedback.instance_id == inst_id)
    )).scalar_one_or_none()
    if not fb:
        raise HTTPException(404, "Feedback not found")
    actions = list(fb.required_actions or [])
    if action_idx < 0 or action_idx >= len(actions):
        raise HTTPException(400, f"Action index {action_idx} out of range (0-{len(actions) - 1})")
    actions[action_idx] = {**actions[action_idx], "addressed": data.addressed}
    fb.required_actions = actions
    await db.flush()
    es, esl = await _get_entity_score(db, inst_id, fb.node_id)
    return _feedback_resp(fb, es, esl)


# ============ EVIDENCE ============

ALLOWED_EXTENSIONS = {".pdf", ".docx", ".xlsx", ".pptx", ".png", ".jpg", ".jpeg", ".zip"}
MAX_FILE_SIZE = 50 * 1024 * 1024

@router.post("/api/assessments/{inst_id}/regulator-feedback/{feedback_id}/evidence", status_code=201)
async def upload_regulator_evidence(
    inst_id: uuid.UUID, feedback_id: uuid.UUID, file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db), current_user: User = Depends(require_role("admin")),
):
    fb = (await db.execute(
        select(RegulatorFeedback).where(RegulatorFeedback.id == feedback_id, RegulatorFeedback.instance_id == inst_id)
    )).scalar_one_or_none()
    if not fb:
        raise HTTPException(404, "Feedback not found")
    filename = file.filename or "unnamed"
    ext = os.path.splitext(filename)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(400, f"File type '{ext}' not allowed. Accepted: {', '.join(ALLOWED_EXTENSIONS)}")
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(400, "File exceeds maximum size of 50MB")
    dir_path = os.path.join(REGULATOR_EVIDENCE_DIR, str(inst_id), str(feedback_id))
    os.makedirs(dir_path, exist_ok=True)
    file_path = os.path.join(dir_path, filename)
    async with aiofiles.open(file_path, "wb") as f:
        await f.write(content)
    ev = RegulatorEvidence(
        feedback_id=feedback_id, file_name=filename, file_path=file_path,
        file_size=len(content), file_type=file.content_type, uploaded_by=current_user.id,
    )
    db.add(ev)
    await db.flush()
    return {"id": str(ev.id), "file_name": ev.file_name, "file_size": ev.file_size, "uploaded_at": ev.uploaded_at.isoformat()}


@router.get("/api/assessments/{inst_id}/regulator-feedback/{feedback_id}/evidence")
async def list_regulator_evidence(
    inst_id: uuid.UUID, feedback_id: uuid.UUID,
    db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user),
):
    evs = (await db.execute(
        select(RegulatorEvidence).where(RegulatorEvidence.feedback_id == feedback_id)
    )).scalars().all()
    return [{
        "id": str(e.id), "file_name": e.file_name, "file_size": e.file_size, "file_type": e.file_type,
        "description": e.description, "uploaded_by": str(e.uploaded_by), "uploaded_at": e.uploaded_at.isoformat(),
        "download_url": f"/api/assessments/regulator-evidence/{e.id}/download",
    } for e in evs]


@router.get("/api/assessments/regulator-evidence/{evidence_id}/download")
async def download_regulator_evidence(evidence_id: uuid.UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    ev = (await db.execute(select(RegulatorEvidence).where(RegulatorEvidence.id == evidence_id))).scalar_one_or_none()
    if not ev:
        raise HTTPException(404, "Evidence not found")
    if not os.path.exists(ev.file_path):
        raise HTTPException(404, "File not found on disk")
    return FileResponse(ev.file_path, filename=ev.file_name, media_type=ev.file_type or "application/octet-stream")


@router.delete("/api/assessments/regulator-evidence/{evidence_id}", status_code=204)
async def delete_regulator_evidence(evidence_id: uuid.UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(require_role("admin"))):
    ev = (await db.execute(select(RegulatorEvidence).where(RegulatorEvidence.id == evidence_id))).scalar_one_or_none()
    if not ev:
        raise HTTPException(404, "Evidence not found")
    if os.path.exists(ev.file_path):
        os.remove(ev.file_path)
    await db.delete(ev)
    await db.flush()


# ============ BULK UPLOAD / EXPORT TEMPLATE ============

AGREEMENT_MAP = {
    "agreed": "agreed", "agree": "agreed", "yes": "agreed",
    "disagreed": "disagreed", "disagree": "disagreed", "no": "disagreed",
    "partially_agreed": "partially_agreed", "partially agreed": "partially_agreed", "partial": "partially_agreed", "partially": "partially_agreed",
}
PRIORITY_MAP = {
    "critical": "critical", "major": "major", "minor": "minor", "observation": "observation",
}


@router.get("/api/assessments/{inst_id}/regulator-feedback/export-template")
async def export_feedback_template(
    inst_id: uuid.UUID, phase_id: uuid.UUID | None = None,
    db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user),
):
    inst = (await db.execute(select(AssessmentInstance).where(AssessmentInstance.id == inst_id))).scalar_one_or_none()
    if not inst:
        raise HTTPException(404, "Assessment instance not found")

    # Get all assessable nodes for this framework
    nodes = (await db.execute(
        select(FrameworkNode)
        .where(FrameworkNode.framework_id == inst.framework_id, FrameworkNode.is_assessable == True)
        .order_by(FrameworkNode.sort_order)
    )).scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Regulator Feedback"
    headers = ["Reference Code", "Node Name", "Node Name (Arabic)", "Entity Score", "Entity Score Label",
               "Agreement", "Regulator Score", "Feedback", "Required Actions", "Priority"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    for row_idx, node in enumerate(nodes, 2):
        es, esl = await _get_entity_score(db, inst_id, node.id)
        ws.cell(row=row_idx, column=1, value=node.reference_code or "")
        ws.cell(row=row_idx, column=2, value=node.name)
        ws.cell(row=row_idx, column=3, value=node.name_ar or "")
        ws.cell(row=row_idx, column=4, value=float(es) if es is not None else "")
        ws.cell(row=row_idx, column=5, value=esl or "")
        # Columns 6-10 are for the regulator to fill in

    # Auto-size columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=regulator_feedback_template.xlsx"},
    )


@router.post("/api/assessments/{inst_id}/regulator-feedback/bulk-upload")
async def bulk_upload_feedback(
    inst_id: uuid.UUID, phase_id: uuid.UUID = Query(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db), current_user: User = Depends(require_role("admin")),
):
    inst = (await db.execute(select(AssessmentInstance).where(AssessmentInstance.id == inst_id))).scalar_one_or_none()
    if not inst:
        raise HTTPException(404, "Assessment instance not found")

    content = await file.read()
    try:
        wb = load_workbook(BytesIO(content))
    except Exception:
        raise HTTPException(400, "Could not parse file. Please upload a valid .xlsx file.")
    ws = wb.active

    # Build node lookup by reference_code
    nodes = (await db.execute(
        select(FrameworkNode)
        .where(FrameworkNode.framework_id == inst.framework_id, FrameworkNode.is_assessable == True)
    )).scalars().all()
    node_map = {n.reference_code: n for n in nodes if n.reference_code}

    # Read headers
    headers = [str(cell.value or "").strip().lower() for cell in ws[1]]
    col_idx = {}
    for i, h in enumerate(headers):
        if "reference" in h and "code" in h: col_idx["ref"] = i
        elif h in ("agreement", "agreement status"): col_idx["agreement"] = i
        elif "regulator" in h and "score" in h: col_idx["reg_score"] = i
        elif h in ("feedback", "feedback text", "comments"): col_idx["feedback"] = i
        elif "required" in h and "action" in h: col_idx["actions"] = i
        elif h == "priority": col_idx["priority"] = i

    if "ref" not in col_idx:
        raise HTTPException(400, "Missing 'Reference Code' column in the uploaded file.")

    imported = 0
    skipped = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[col_idx["ref"]]:
            skipped += 1
            continue
        ref_code = str(row[col_idx["ref"]]).strip()
        node = node_map.get(ref_code)
        if not node:
            errors.append({"row": row_idx, "error": f"Node not found: '{ref_code}'"})
            skipped += 1
            continue

        # Parse fields
        agreement_raw = str(row[col_idx.get("agreement", -1)] or "").strip().lower() if "agreement" in col_idx else ""
        agreement = AGREEMENT_MAP.get(agreement_raw, "not_reviewed")

        reg_score = None
        if "reg_score" in col_idx and row[col_idx["reg_score"]] is not None:
            try:
                reg_score = Decimal(str(row[col_idx["reg_score"]]))
            except Exception:
                pass

        feedback_text = str(row[col_idx.get("feedback", -1)] or "").strip() if "feedback" in col_idx else ""
        priority_raw = str(row[col_idx.get("priority", -1)] or "").strip().lower() if "priority" in col_idx else ""
        priority = PRIORITY_MAP.get(priority_raw, "observation")

        # Parse required actions (semicolon-separated)
        actions = []
        if "actions" in col_idx and row[col_idx["actions"]]:
            raw_actions = str(row[col_idx["actions"]]).split(";")
            for a in raw_actions:
                a = a.strip()
                if a:
                    actions.append({"action": a, "priority": priority, "addressed": False})

        # Resolve score label
        score_label = None
        if reg_score is not None:
            score_label = await _resolve_score_label(db, inst.framework_id, float(reg_score))

        # Upsert
        existing = (await db.execute(
            select(RegulatorFeedback).where(
                RegulatorFeedback.instance_id == inst_id,
                RegulatorFeedback.node_id == node.id,
                RegulatorFeedback.phase_id == phase_id,
            )
        )).scalar_one_or_none()

        if existing:
            existing.agreement_status = agreement
            existing.regulator_score = reg_score
            existing.regulator_score_label = score_label
            existing.feedback_text = feedback_text or existing.feedback_text
            existing.required_actions = actions if actions else existing.required_actions
            existing.priority = priority
            existing.updated_by = current_user.id
        else:
            db.add(RegulatorFeedback(
                instance_id=inst_id, node_id=node.id, phase_id=phase_id,
                agreement_status=agreement, regulator_score=reg_score,
                regulator_score_label=score_label, feedback_text=feedback_text or None,
                required_actions=actions, priority=priority, feedback_by=current_user.id,
            ))
        imported += 1

    await db.flush()
    return {"total_rows": row_idx - 1, "imported": imported, "skipped": skipped, "errors": errors}


# ============ SUMMARY & STATS ============

@router.get("/api/assessments/{inst_id}/regulator-feedback/summary")
async def feedback_summary(
    inst_id: uuid.UUID, phase_id: uuid.UUID | None = None,
    db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user),
):
    inst = (await db.execute(select(AssessmentInstance).where(AssessmentInstance.id == inst_id))).scalar_one_or_none()
    if not inst:
        raise HTTPException(404, "Assessment instance not found")

    # Total assessable nodes
    total_nodes = (await db.execute(
        select(func.count()).select_from(FrameworkNode)
        .where(FrameworkNode.framework_id == inst.framework_id, FrameworkNode.is_assessable == True)
    )).scalar() or 0

    base = select(RegulatorFeedback).where(RegulatorFeedback.instance_id == inst_id)
    if phase_id:
        base = base.where(RegulatorFeedback.phase_id == phase_id)

    feedbacks = (await db.execute(base)).scalars().all()
    reviewed = len(feedbacks)

    # Agreement breakdown
    agreed = sum(1 for f in feedbacks if f.agreement_status == "agreed")
    disagreed = sum(1 for f in feedbacks if f.agreement_status == "disagreed")
    partially = sum(1 for f in feedbacks if f.agreement_status == "partially_agreed")
    not_reviewed_count = total_nodes - reviewed

    # Priority breakdown
    by_priority = {"critical": 0, "major": 0, "minor": 0, "observation": 0}
    for f in feedbacks:
        if f.priority in by_priority:
            by_priority[f.priority] += 1

    # Score averages
    entity_scores = []
    reg_scores = []
    for f in feedbacks:
        if f.regulator_score is not None:
            reg_scores.append(float(f.regulator_score))
        es, _ = await _get_entity_score(db, inst_id, f.node_id)
        if es is not None:
            entity_scores.append(float(es))

    avg_entity = round(sum(entity_scores) / len(entity_scores), 2) if entity_scores else None
    avg_reg = round(sum(reg_scores) / len(reg_scores), 2) if reg_scores else None
    avg_gap = round(avg_entity - avg_reg, 2) if avg_entity is not None and avg_reg is not None else None

    # Correction breakdown
    corrections = {"pending": 0, "in_progress": 0, "addressed": 0, "accepted": 0, "rejected": 0}
    for f in feedbacks:
        if f.agreement_status in ("disagreed", "partially_agreed") and f.correction_status in corrections:
            corrections[f.correction_status] += 1

    return {
        "total_nodes": total_nodes,
        "reviewed_nodes": reviewed,
        "not_reviewed": not_reviewed_count,
        "agreed": agreed,
        "disagreed": disagreed,
        "partially_agreed": partially,
        "by_priority": by_priority,
        "avg_entity_score": avg_entity,
        "avg_regulator_score": avg_reg,
        "avg_gap": avg_gap,
        "corrections": corrections,
    }


@router.get("/api/assessments/{inst_id}/score-comparison")
async def score_comparison(
    inst_id: uuid.UUID, phase_id: uuid.UUID | None = None,
    db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user),
):
    inst = (await db.execute(select(AssessmentInstance).where(AssessmentInstance.id == inst_id))).scalar_one_or_none()
    if not inst:
        raise HTTPException(404, "Assessment instance not found")

    # All assessable nodes
    nodes = (await db.execute(
        select(FrameworkNode)
        .where(FrameworkNode.framework_id == inst.framework_id, FrameworkNode.is_assessable == True)
        .order_by(FrameworkNode.sort_order)
    )).scalars().all()

    # Get all feedback indexed by node_id
    fb_query = select(RegulatorFeedback).where(RegulatorFeedback.instance_id == inst_id)
    if phase_id:
        fb_query = fb_query.where(RegulatorFeedback.phase_id == phase_id)
    feedbacks = (await db.execute(fb_query)).scalars().all()
    fb_map = {fb.node_id: fb for fb in feedbacks}

    results = []
    for node in nodes:
        es, esl = await _get_entity_score(db, inst_id, node.id)
        fb = fb_map.get(node.id)
        reg_score = float(fb.regulator_score) if fb and fb.regulator_score is not None else None
        entity_score = float(es) if es is not None else None
        gap = round(entity_score - reg_score, 2) if entity_score is not None and reg_score is not None else None

        results.append({
            "node_id": str(node.id),
            "reference_code": node.reference_code,
            "name": node.name,
            "name_ar": node.name_ar,
            "node_type": node.node_type,
            "parent_id": str(node.parent_id) if node.parent_id else None,
            "entity_score": entity_score,
            "entity_score_label": esl,
            "regulator_score": reg_score,
            "regulator_score_label": fb.regulator_score_label if fb else None,
            "gap": gap,
            "agreement_status": fb.agreement_status if fb else "not_reviewed",
            "priority": fb.priority if fb else None,
            "correction_status": fb.correction_status if fb else None,
        })
    return results
