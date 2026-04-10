import uuid
from io import BytesIO

from sqlalchemy.ext.asyncio import AsyncSession

from app.data.domains import DOMAINS
from app.services.dashboard_service import get_entity_dashboard, get_completion_matrix


async def export_excel(db: AsyncSession, entity_id: uuid.UUID) -> BytesIO:
    """Export entity completion matrix as Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise RuntimeError("openpyxl not installed")

    dashboard = await get_entity_dashboard(db, entity_id)
    if not dashboard:
        raise ValueError("Entity not found")

    matrix = await get_completion_matrix(db, entity_id)

    wb = openpyxl.Workbook()

    # Sheet 1: Entity Info
    ws1 = wb.active
    ws1.title = "Entity Overview"
    header_font = Font(bold=True, size=12)
    ws1["A1"] = "Entity Name (EN)"
    ws1["B1"] = dashboard["name_en"]
    ws1["A2"] = "Entity Name (AR)"
    ws1["B2"] = dashboard["name_ar"]
    ws1["A3"] = "Sector"
    ws1["B3"] = dashboard.get("sector", "")
    ws1["A4"] = "Total Products"
    ws1["B4"] = dashboard["product_count"]
    ws1["A5"] = "Completion"
    ws1["B5"] = f"{dashboard['completion_percentage']}%"
    ws1["A6"] = "Badge Tier"
    ws1["B6"] = dashboard.get("badge_tier", "Not assigned")
    for row in ws1.iter_rows(min_row=1, max_row=6, min_col=1, max_col=1):
        for cell in row:
            cell.font = Font(bold=True)

    # Sheet 2: Completion Matrix
    ws2 = wb.create_sheet("Completion Matrix")
    fill_complete = PatternFill(start_color="C6EFCE", fill_type="solid")
    fill_progress = PatternFill(start_color="FFEB9C", fill_type="solid")
    fill_not_started = PatternFill(start_color="F2F2F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Headers
    ws2.cell(row=1, column=1, value="Product").font = Font(bold=True)
    for did in range(1, 12):
        cell = ws2.cell(row=1, column=did + 1, value=f"D{did}")
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data rows
    for i, row in enumerate(matrix, start=2):
        ws2.cell(row=i, column=1, value=row["product_name"] or "Untitled").border = thin_border
        for did in range(1, 12):
            status = row["domains"].get(did, "not_started")
            cell = ws2.cell(row=i, column=did + 1, value=status.replace("_", " ").title())
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            cell.font = Font(size=9)
            if status == "complete":
                cell.fill = fill_complete
            elif status == "in_progress":
                cell.fill = fill_progress
            else:
                cell.fill = fill_not_started

    ws2.column_dimensions["A"].width = 25
    for did in range(1, 12):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(did + 1)].width = 14

    # Sheet 3: Product Details
    ws3 = wb.create_sheet("Product Details")
    ws3.cell(row=1, column=1, value="Product").font = Font(bold=True)
    ws3.cell(row=1, column=2, value="Status").font = Font(bold=True)
    ws3.cell(row=1, column=3, value="Filled Sub-Reqs").font = Font(bold=True)
    ws3.cell(row=1, column=4, value="Total Sub-Reqs").font = Font(bold=True)
    ws3.cell(row=1, column=5, value="Completion %").font = Font(bold=True)

    for i, p in enumerate(dashboard["products"], start=2):
        ws3.cell(row=i, column=1, value=p["name_en"] or "Untitled")
        ws3.cell(row=i, column=2, value=p["status"].title())
        ws3.cell(row=i, column=3, value=p["filled_count"])
        ws3.cell(row=i, column=4, value=p["total_count"])
        pct = (p["filled_count"] / p["total_count"] * 100) if p["total_count"] > 0 else 0
        ws3.cell(row=i, column=5, value=f"{pct:.1f}%")

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
