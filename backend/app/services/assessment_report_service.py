"""Assessment Report Export Service — generates multi-sheet Excel report."""
import uuid
from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.assessment_engine import (
    AssessmentInstance, AssessmentResponse, AssessmentEvidence,
    AssessmentResponseHistory, AssessedEntity, AiProduct,
)
from app.models.compliance_framework import ComplianceFramework
from app.models.framework_node import FrameworkNode

# Styles
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="00338D", end_color="00338D", fill_type="solid")
COMPLIANT_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
SEMI_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
NON_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"), bottom=Side(style="thin", color="D1D5DB"),
)


def _style_header(ws, cols):
    for col_idx, (title, width) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = width


def _style_cell(cell, wrap=False):
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)


async def generate_assessment_report(db: AsyncSession, instance_id: uuid.UUID) -> BytesIO:
    # Load instance
    inst = (await db.execute(
        select(AssessmentInstance).options(
            selectinload(AssessmentInstance.cycle), selectinload(AssessmentInstance.framework),
            selectinload(AssessmentInstance.assessed_entity),
        ).where(AssessmentInstance.id == instance_id)
    )).scalar_one_or_none()
    if not inst:
        raise ValueError("Assessment instance not found")

    # Load responses with nodes
    responses = (await db.execute(
        select(AssessmentResponse).options(selectinload(AssessmentResponse.node))
        .where(AssessmentResponse.instance_id == instance_id)
    )).scalars().all()

    # Load products
    product_ids = list({r.ai_product_id for r in responses if r.ai_product_id})
    products = {}
    if product_ids:
        prods = (await db.execute(select(AiProduct).where(AiProduct.id.in_(product_ids)))).scalars().all()
        products = {p.id: p for p in prods}

    # Load all evidence
    resp_ids = [r.id for r in responses]
    all_evidence = []
    if resp_ids:
        all_evidence = (await db.execute(select(AssessmentEvidence).where(AssessmentEvidence.response_id.in_(resp_ids)))).scalars().all()
    evidence_by_resp = {}
    for ev in all_evidence:
        evidence_by_resp.setdefault(ev.response_id, []).append(ev)

    # Load history
    all_history = []
    if resp_ids:
        all_history = (await db.execute(
            select(AssessmentResponseHistory).where(AssessmentResponseHistory.response_id.in_(resp_ids))
            .order_by(AssessmentResponseHistory.changed_at.asc())
        )).scalars().all()

    # Node map for reference codes
    resp_node_map = {r.id: r for r in responses}

    wb = Workbook()

    # ========== Sheet 1: Summary ==========
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50

    rows = [
        ("Entity", inst.assessed_entity.name if inst.assessed_entity else ""),
        ("Framework", f"{inst.framework.name} ({inst.framework.abbreviation})" if inst.framework else ""),
        ("Cycle", inst.cycle.cycle_name if inst.cycle else ""),
        ("Status", inst.status),
        ("Started", inst.started_at.strftime("%Y-%m-%d %H:%M") if inst.started_at else "Not started"),
        ("Submitted", inst.submitted_at.strftime("%Y-%m-%d %H:%M") if inst.submitted_at else ""),
        ("Completed", inst.completed_at.strftime("%Y-%m-%d %H:%M") if inst.completed_at else ""),
        ("", ""),
        ("Total Assessable Nodes", inst.total_assessable_nodes),
        ("Answered", inst.answered_nodes),
        ("Progress", f"{round((inst.answered_nodes / inst.total_assessable_nodes) * 100)}%" if inst.total_assessable_nodes else "0%"),
    ]

    # Compliance stats
    compliant = sum(1 for r in responses if r.computed_score_label == "Compliant")
    semi = sum(1 for r in responses if r.computed_score_label == "Semi-Compliant")
    non_c = sum(1 for r in responses if r.computed_score_label == "Non-Compliant")
    pending = sum(1 for r in responses if r.status == "pending")

    rows += [
        ("", ""),
        ("Compliant", compliant),
        ("Semi-Compliant", semi),
        ("Non-Compliant", non_c),
        ("Pending", pending),
        ("Evidence Files", len(all_evidence)),
    ]

    # Products
    if products:
        rows.append(("", ""))
        rows.append(("AI Products", len(products)))
        for p in products.values():
            prod_resps = [r for r in responses if r.ai_product_id == p.id]
            pc = sum(1 for r in prod_resps if r.computed_score_label == "Compliant")
            ps = sum(1 for r in prod_resps if r.computed_score_label == "Semi-Compliant")
            pn = sum(1 for r in prod_resps if r.computed_score_label == "Non-Compliant")
            total_answered = pc + ps + pn
            pct = round((pc / total_answered) * 100) if total_answered else 0
            rows.append((f"  {p.name}", f"{pct}% compliant ({pc}C / {ps}SC / {pn}NC)"))

    for row_idx, (label, value) in enumerate(rows, 1):
        c1 = ws.cell(row=row_idx, column=1, value=label)
        c1.font = Font(bold=True) if label else Font()
        ws.cell(row=row_idx, column=2, value=value)

    # ========== Sheet 2: Node Details ==========
    ws2 = wb.create_sheet("Node Details")
    cols2 = [("Reference", 12), ("Node Name", 35), ("Product", 25), ("Compliance", 16), ("Answer", 50),
             ("Review Feedback", 45), ("Justification", 45), ("Recommendations", 45), ("Status", 12)]
    _style_header(ws2, cols2)

    sorted_resps = sorted(responses, key=lambda r: (r.node.reference_code or "") if r.node else "")
    for row_idx, r in enumerate(sorted_resps, 2):
        rd = r.response_data or {}
        prod_name = products[r.ai_product_id].name if r.ai_product_id and r.ai_product_id in products else ""
        vals = [
            r.node.reference_code if r.node else "", r.node.name if r.node else "", prod_name,
            r.computed_score_label or "", rd.get("answer", ""), rd.get("review_feedback", ""),
            rd.get("justification", ""), rd.get("recommendation", ""), r.status,
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            _style_cell(cell, wrap=col_idx >= 5)
        # Color compliance
        label = r.computed_score_label
        if label == "Compliant":
            for c in range(1, len(vals) + 1): ws2.cell(row=row_idx, column=c).fill = COMPLIANT_FILL
        elif label == "Semi-Compliant":
            for c in range(1, len(vals) + 1): ws2.cell(row=row_idx, column=c).fill = SEMI_FILL
        elif label == "Non-Compliant":
            for c in range(1, len(vals) + 1): ws2.cell(row=row_idx, column=c).fill = NON_FILL

    # ========== Sheet 3: Evidence Log ==========
    ws3 = wb.create_sheet("Evidence Log")
    cols3 = [("Node Ref", 12), ("Node Name", 30), ("Product", 25), ("File Name", 35), ("Type", 15),
             ("Size (KB)", 12), ("Doc Date", 12), ("Approved", 10), ("Approved By", 20),
             ("Signature", 10), ("Reviewer Notes", 40), ("Uploaded", 18)]
    _style_header(ws3, cols3)

    row_idx = 2
    for r in sorted_resps:
        evs = evidence_by_resp.get(r.id, [])
        prod_name = products[r.ai_product_id].name if r.ai_product_id and r.ai_product_id in products else ""
        for ev in evs:
            vals = [
                r.node.reference_code if r.node else "", r.node.name if r.node else "", prod_name,
                ev.file_name, ev.file_type or "", round((ev.file_size or 0) / 1024, 1),
                ev.document_date.isoformat() if ev.document_date else "",
                "Yes" if ev.is_approved else ("No" if ev.is_approved is False else ""),
                ev.approved_by or "", "Yes" if ev.has_signature else "",
                ev.reviewer_notes or "", ev.uploaded_at.strftime("%Y-%m-%d %H:%M") if ev.uploaded_at else "",
            ]
            for col_idx, val in enumerate(vals, 1):
                cell = ws3.cell(row=row_idx, column=col_idx, value=val)
                _style_cell(cell, wrap=col_idx == 11)
            row_idx += 1

    # ========== Sheet 4: Review History ==========
    ws4 = wb.create_sheet("Review History")
    cols4 = [("Node Ref", 12), ("Node Name", 30), ("Round", 8), ("Compliance", 16),
             ("Reviewer Feedback", 50), ("Changed By", 20), ("Changed At", 18), ("Change Type", 14)]
    _style_header(ws4, cols4)

    row_idx = 2
    for h in all_history:
        resp = resp_node_map.get(h.response_id)
        node = resp.node if resp else None
        vals = [
            node.reference_code if node else "", node.name if node else "",
            h.review_round or "", h.computed_score_label or "",
            h.reviewer_feedback or "", str(h.changed_by) if h.changed_by else "",
            h.changed_at.strftime("%Y-%m-%d %H:%M") if h.changed_at else "", h.change_type,
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=val)
            _style_cell(cell, wrap=col_idx == 5)
        row_idx += 1

    # ========== Sheet 5: Gap Analysis ==========
    ws5 = wb.create_sheet("Gap Analysis")
    cols5 = [("Reference", 12), ("Node Name", 35), ("Product", 25), ("Status", 16),
             ("Review Feedback", 50), ("Recommendations", 50), ("Evidence Count", 14)]
    _style_header(ws5, cols5)

    row_idx = 2
    gaps = [r for r in sorted_resps if r.computed_score_label in ("Semi-Compliant", "Non-Compliant")]
    for r in gaps:
        rd = r.response_data or {}
        prod_name = products[r.ai_product_id].name if r.ai_product_id and r.ai_product_id in products else ""
        ev_count = len(evidence_by_resp.get(r.id, []))
        vals = [
            r.node.reference_code if r.node else "", r.node.name if r.node else "", prod_name,
            r.computed_score_label or "", rd.get("review_feedback", ""),
            rd.get("recommendation", ""), ev_count,
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws5.cell(row=row_idx, column=col_idx, value=val)
            _style_cell(cell, wrap=col_idx >= 5)
        fill = SEMI_FILL if r.computed_score_label == "Semi-Compliant" else NON_FILL
        for c in range(1, len(vals) + 1): ws5.cell(row=row_idx, column=c).fill = fill
        row_idx += 1

    # Save
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
