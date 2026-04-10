"""
Seed NDI hierarchy from NDI_Domains_Controls_Data.xlsx.
Replaces existing NDI hierarchy with fresh data from 5 sheets:
- Domains (14 nodes, depth 0)
- Maturity Questions (42 nodes, depth 1)
- Specifications (24 nodes, depth 2)
Also stores maturity levels + acceptance criteria in metadata_json.
"""
import asyncio, uuid, sys, os, json
from decimal import Decimal
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from sqlalchemy import select, delete
from app.database import async_session
from app.models.compliance_framework import ComplianceFramework
from app.models.framework_node import FrameworkNode, NodeType

EXCEL = "/app/NDI_Domains_Controls_Data.xlsx"

async def main():
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    print(f"Sheets: {wb.sheetnames}")

    async with async_session() as db:
        fw = (await db.execute(select(ComplianceFramework).where(ComplianceFramework.abbreviation == "NDI"))).scalar_one()
        print(f"NDI framework: {fw.id}")

        # Clear existing nodes
        await db.execute(delete(FrameworkNode).where(FrameworkNode.framework_id == fw.id))
        await db.flush()
        print("Cleared existing NDI nodes")

        # Ensure node types exist
        existing_types = {nt.name for nt in (await db.execute(select(NodeType).where(NodeType.framework_id == fw.id))).scalars().all()}
        for name, color, assessable in [("Domain", "#0091DA", False), ("Question", "#E67E22", True), ("Specification", "#27AE60", True)]:
            if name not in existing_types:
                db.add(NodeType(framework_id=fw.id, name=name, label=name, color=color, is_assessable_default=assessable))
                print(f"  Added node type: {name}")
        await db.flush()

        # ========== Read Maturity Levels (per question) ==========
        ws_levels = wb["Maturity Levels"]
        # question_code -> [{level, name_en, name_ar, desc_en, desc_ar, evidence_en, evidence_ar}]
        question_levels = {}
        for row in ws_levels.iter_rows(min_row=2, max_row=ws_levels.max_row, values_only=True):
            if not row[1]:
                continue
            qcode = str(row[1]).strip()
            level = int(row[2]) if row[2] is not None else 0
            if qcode not in question_levels:
                question_levels[qcode] = []
            question_levels[qcode].append({
                "level": level,
                "name_en": str(row[3]).strip() if row[3] else "",
                "name_ar": str(row[4]).strip() if row[4] else "",
                "description_en": str(row[5]).strip() if row[5] else "",
                "description_ar": str(row[6]).strip() if row[6] else "",
                "evidence_en": str(row[7]).strip() if row[7] else "",
                "evidence_ar": str(row[8]).strip() if row[8] else "",
            })
        print(f"Loaded maturity levels for {len(question_levels)} questions")

        # ========== Read Acceptance Criteria (per question per level) ==========
        ws_criteria = wb["Acceptance Criteria"]
        # question_code -> {level -> [criteria]}
        question_criteria = {}
        for row in ws_criteria.iter_rows(min_row=2, max_row=ws_criteria.max_row, values_only=True):
            if not row[1]:
                continue
            qcode = str(row[1]).strip()
            level = int(row[2]) if row[2] is not None else 0
            if qcode not in question_criteria:
                question_criteria[qcode] = {}
            if level not in question_criteria[qcode]:
                question_criteria[qcode][level] = []
            question_criteria[qcode][level].append({
                "text_en": str(row[4]).strip() if row[4] else "",
                "text_ar": str(row[5]).strip() if row[5] else "",
                "order": int(row[6]) if row[6] else 0,
            })
        print(f"Loaded acceptance criteria for {len(question_criteria)} questions")

        ref_to_id = {}
        ref_to_path = {}
        count = 0

        # ========== Sheet 1: Domains ==========
        ws_domains = wb["Domains"]
        for row in ws_domains.iter_rows(min_row=2, max_row=ws_domains.max_row, values_only=True):
            code = str(row[0]).strip() if row[0] else None
            if not code:
                continue
            name_en = str(row[1]).strip() if row[1] else code
            name_ar = str(row[2]).strip() if row[2] else None
            has_oe = str(row[3]).strip() if row[3] else "No"
            weight = float(row[4]) if row[4] else None
            sort_order = int(row[6]) if row[6] else count

            node_id = uuid.uuid4()
            path = f"/{node_id}/"
            db.add(FrameworkNode(
                id=node_id, framework_id=fw.id, parent_id=None,
                node_type="Domain", reference_code=code,
                name=name_en, name_ar=name_ar,
                sort_order=sort_order, path=path, depth=0,
                is_active=True, is_assessable=False,
                weight=Decimal(str(weight)) if weight else None,
                metadata_json={"has_operational_excellence": has_oe == "Yes"},
            ))
            ref_to_id[code] = node_id
            ref_to_path[code] = path
            count += 1

        await db.flush()
        print(f"Inserted {count} domains")

        # ========== Sheet 2: Maturity Questions ==========
        ws_questions = wb["Maturity Questions"]
        q_count = 0
        for row in ws_questions.iter_rows(min_row=2, max_row=ws_questions.max_row, values_only=True):
            domain_code = str(row[0]).strip() if row[0] else None
            qcode = str(row[2]).strip() if row[2] else None
            if not qcode or not domain_code:
                continue
            question_en = str(row[3]).strip() if row[3] else qcode
            question_ar = str(row[4]).strip() if row[4] else None
            sort_order = int(row[5]) if row[5] else q_count

            parent_id = ref_to_id.get(domain_code)
            if not parent_id:
                print(f"  SKIP question {qcode}: parent {domain_code} not found")
                continue

            node_id = uuid.uuid4()
            parent_path = ref_to_path.get(domain_code, "/")
            path = f"{parent_path}{node_id}/"

            # Build metadata with maturity levels + acceptance criteria
            metadata = {}
            if qcode in question_levels:
                metadata["maturity_levels"] = question_levels[qcode]
            if qcode in question_criteria:
                criteria_by_level = {}
                for lvl, criteria_list in question_criteria[qcode].items():
                    criteria_by_level[str(lvl)] = sorted(criteria_list, key=lambda x: x["order"])
                metadata["acceptance_criteria"] = criteria_by_level

            db.add(FrameworkNode(
                id=node_id, framework_id=fw.id, parent_id=parent_id,
                node_type="Question", reference_code=qcode,
                name=question_en, name_ar=question_ar,
                sort_order=sort_order, path=path, depth=1,
                is_active=True, is_assessable=True,
                metadata_json=metadata if metadata else None,
            ))
            ref_to_id[qcode] = node_id
            ref_to_path[qcode] = path
            q_count += 1

        await db.flush()
        print(f"Inserted {q_count} maturity questions (with levels + criteria in metadata)")

        # ========== Sheet 5: Specifications ==========
        ws_specs = wb["Specifications"]
        s_count = 0
        for row in ws_specs.iter_rows(min_row=2, max_row=ws_specs.max_row, values_only=True):
            spec_code = str(row[0]).strip() if row[0] else None
            qcode = str(row[2]).strip() if row[2] else None
            if not spec_code or not qcode:
                continue
            level = int(row[3]) if row[3] is not None else 0
            desc_en = str(row[4]).strip() if row[4] else spec_code
            desc_ar = str(row[5]).strip() if row[5] else None

            parent_id = ref_to_id.get(qcode)
            if not parent_id:
                print(f"  SKIP spec {spec_code}: parent {qcode} not found")
                continue

            node_id = uuid.uuid4()
            parent_path = ref_to_path.get(qcode, "/")
            path = f"{parent_path}{node_id}/"

            db.add(FrameworkNode(
                id=node_id, framework_id=fw.id, parent_id=parent_id,
                node_type="Specification", reference_code=spec_code,
                name=desc_en, name_ar=desc_ar,
                sort_order=s_count, path=path, depth=2,
                is_active=True, is_assessable=True,
                weight=Decimal("1.0"), max_score=Decimal("1.0"),
                metadata_json={"maturity_level": level},
            ))
            ref_to_id[spec_code] = node_id
            ref_to_path[spec_code] = path
            s_count += 1

        await db.flush()
        await db.commit()
        total = count + q_count + s_count
        print(f"\nDone! Total: {total} nodes ({count} domains + {q_count} questions + {s_count} specifications)")
        print(f"Maturity levels stored in metadata for {len(question_levels)} questions")
        print(f"Acceptance criteria stored for {len(question_criteria)} questions ({sum(sum(len(v) for v in q.values()) for q in question_criteria.values())} total criteria)")

if __name__ == "__main__":
    asyncio.run(main())
