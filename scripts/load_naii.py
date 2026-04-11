#!/usr/bin/env python3
"""Load NAII framework hierarchy from Excel into the database."""
import sys
import io
import uuid

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

EXCEL_PATH = r'C:\Users\amroa\Downloads\NAII_Framework_Complete.xlsx'
FW_ID = '3a4b98bc-fb87-41d9-bcb1-caddced27de5'
SQL_OUTPUT = r'C:\projects\AI-Badges\scripts\naii_insert.sql'

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['NAII Maturity Questions']

# Parse all data rows
rows_data = []
for row in range(2, ws.max_row + 1):
    code = ws.cell(row=row, column=1).value
    if not code:
        continue
    rows_data.append({
        'code': str(code).strip(),
        'domain_en': str(ws.cell(row=row, column=2).value or '').strip(),
        'domain_ar': str(ws.cell(row=row, column=3).value or '').strip(),
        'subdomain_ar': str(ws.cell(row=row, column=4).value or '').strip(),
        'subdomain_en': str(ws.cell(row=row, column=5).value or '').strip(),
        'question_ar': str(ws.cell(row=row, column=6).value or '').strip(),
        'level': ws.cell(row=row, column=7).value,
        'level_name_ar': str(ws.cell(row=row, column=8).value or '').strip(),
        'level_name_en': str(ws.cell(row=row, column=9).value or '').strip(),
        'level_desc_ar': str(ws.cell(row=row, column=10).value or '').strip(),
        'evidence_ar': str(ws.cell(row=row, column=11).value or '').strip(),
    })

# Build hierarchy structures
domains = {}
subdomains = {}
questions = {}

domain_sort = 0
subdomain_sort = 0
question_sort = 0

for r in rows_data:
    # Domains
    if r['domain_en'] not in domains:
        domain_sort += 1
        code_parts = r['code'].split('.')
        domain_abbr = code_parts[2] if len(code_parts) > 2 else f'D{domain_sort}'
        domains[r['domain_en']] = {
            'en': r['domain_en'],
            'ar': r['domain_ar'],
            'ref': f'D.{domain_abbr}',
            'sort': domain_sort,
        }

    # Subdomains
    sd_key = f"{r['domain_en']}|{r['subdomain_en']}"
    if sd_key not in subdomains:
        subdomain_sort += 1
        # Count how many subdomains under this domain so far
        domain_sd_count = sum(1 for sk in subdomains if sk.startswith(r['domain_en'] + '|'))
        domain_abbr = domains[r['domain_en']]['ref'].split('.')[-1]  # ST or GO
        sd_ref = f"SD.{domain_abbr}.{domain_sd_count + 1}"
        subdomains[sd_key] = {
            'en': r['subdomain_en'],
            'ar': r['subdomain_ar'],
            'domain_en': r['domain_en'],
            'sort': subdomain_sort,
            'ref': sd_ref,
        }

    # Questions (group by code)
    if r['code'] not in questions:
        question_sort += 1
        questions[r['code']] = {
            'code': r['code'],
            'question_ar': r['question_ar'],
            'domain_en': r['domain_en'],
            'subdomain_en': r['subdomain_en'],
            'sort': question_sort,
            'levels': [],
        }
    questions[r['code']]['levels'].append({
        'level': r['level'],
        'name_ar': r['level_name_ar'],
        'name_en': r['level_name_en'],
        'desc_ar': r['level_desc_ar'],
        'evidence_ar': r['evidence_ar'],
    })


def esc(s):
    """Escape single quotes for SQL."""
    return s.replace("'", "''")


sql_lines = []

# Create domain nodes (depth=0, path=ref)
domain_ids = {}
domain_refs = {}
for d in domains.values():
    did = str(uuid.uuid4())
    domain_ids[d['en']] = did
    domain_refs[d['en']] = d['ref']
    sql_lines.append(
        f"INSERT INTO framework_nodes "
        f"(id, framework_id, parent_id, name, name_ar, reference_code, node_type, "
        f"is_assessable, is_active, sort_order, path, depth, created_at, updated_at) "
        f"VALUES ('{did}', '{FW_ID}', NULL, '{esc(d['en'])}', '{esc(d['ar'])}', "
        f"'{d['ref']}', 'Domain', false, true, {d['sort']}, "
        f"'{d['ref']}', 0, NOW(), NOW());"
    )

# Create subdomain nodes (depth=1, path=parent_ref/ref)
subdomain_ids = {}
subdomain_refs = {}
for sd_key, sd in subdomains.items():
    sid = str(uuid.uuid4())
    subdomain_ids[sd_key] = sid
    parent_id = domain_ids[sd['domain_en']]
    parent_ref = domain_refs[sd['domain_en']]

    ref = sd['ref']
    subdomain_refs[sd_key] = ref
    path = f"{parent_ref}/{ref}"

    sql_lines.append(
        f"INSERT INTO framework_nodes "
        f"(id, framework_id, parent_id, name, name_ar, reference_code, node_type, "
        f"is_assessable, is_active, sort_order, path, depth, created_at, updated_at) "
        f"VALUES ('{sid}', '{FW_ID}', '{parent_id}', '{esc(sd['en'])}', '{esc(sd['ar'])}', "
        f"'{ref}', 'Sub-Domain', false, true, {sd['sort']}, "
        f"'{path}', 1, NOW(), NOW());"
    )

# Create question nodes (depth=2, path=domain_ref/subdomain_ref/question_ref)
for q in questions.values():
    qid = str(uuid.uuid4())
    sd_key = f"{q['domain_en']}|{q['subdomain_en']}"
    parent_id = subdomain_ids[sd_key]
    parent_ref = subdomain_refs[sd_key]
    domain_ref = domain_refs[q['domain_en']]
    path = f"{domain_ref}/{parent_ref}/{q['code']}"

    # evidence_type from Level Description (AR) - concatenate all levels
    evidence_parts = []
    for lv in sorted(q['levels'], key=lambda x: x['level']):
        evidence_parts.append(
            f"\u0627\u0644\u0645\u0633\u062a\u0648\u0649 {lv['level']} ({lv['name_ar']}): {lv['desc_ar']}"
        )
    evidence_type = '\n'.join(evidence_parts)

    # acceptance_criteria from Supporting Evidence (AR) - concatenate all levels
    criteria_parts = []
    for lv in sorted(q['levels'], key=lambda x: x['level']):
        criteria_parts.append(
            f"\u0627\u0644\u0645\u0633\u062a\u0648\u0649 {lv['level']} ({lv['name_ar']}): {lv['evidence_ar']}"
        )
    acceptance_criteria = '\n'.join(criteria_parts)

    sql_lines.append(
        f"INSERT INTO framework_nodes "
        f"(id, framework_id, parent_id, name, name_ar, reference_code, node_type, "
        f"is_assessable, is_active, sort_order, path, depth, "
        f"evidence_type, acceptance_criteria, created_at, updated_at) "
        f"VALUES ('{qid}', '{FW_ID}', '{parent_id}', '{esc(q['code'])}', "
        f"'{esc(q['question_ar'])}', '{q['code']}', 'Question', true, true, {q['sort']}, "
        f"'{esc(path)}', 2, "
        f"'{esc(evidence_type)}', '{esc(acceptance_criteria)}', NOW(), NOW());"
    )

# Write SQL file
with open(SQL_OUTPUT, 'w', encoding='utf-8') as f:
    f.write('BEGIN;\n')
    for line in sql_lines:
        f.write(line + '\n')
    f.write('COMMIT;\n')

print(f"Generated {len(sql_lines)} SQL inserts:")
print(f"  {len(domains)} domains")
print(f"  {len(subdomains)} sub-domains")
print(f"  {len(questions)} questions")
print(f"SQL written to {SQL_OUTPUT}")
