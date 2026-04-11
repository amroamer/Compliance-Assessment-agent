#!/usr/bin/env python3
"""Load ITGF framework hierarchy from Excel into the database."""
import sys
import io
import uuid
from collections import OrderedDict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

EXCEL_PATH = r'C:\Users\amroa\Downloads\Maste_data_ITGF_12MAR26 - WS1 V1.xlsx'
FW_ID = '24a4c026-b396-400b-85c6-3e1b3d88c670'
SQL_OUTPUT = r'C:\projects\AI-Badges\scripts\itgf_insert.sql'

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

# Build control_name -> control_number map from Map_to_workstream (has actual control numbers)
ws_map = wb['Map_to_workstream']
ctrl_num_map = {}
for row in range(3, ws_map.max_row + 1):
    ctrl_num = ws_map.cell(row=row, column=6).value
    ctrl_name = ws_map.cell(row=row, column=7).value
    if ctrl_num and ctrl_name:
        ctrl_name_clean = str(ctrl_name).strip()
        ctrl_num_map[ctrl_name_clean] = str(ctrl_num).strip()

# Parse Workstream 1_input (header row 3, data from row 4)
ws = wb['Workstream 1_input']

domains = OrderedDict()      # domain_code -> domain_name
subdomains = OrderedDict()   # sd_code -> {name, domain_code}
controls = OrderedDict()     # ctrl_num -> {name, sd_code, domain_code, design, effect}

for row in range(4, ws.max_row + 1):
    domain_code = ws.cell(row=row, column=2).value
    domain_name = ws.cell(row=row, column=3).value
    sd_code = ws.cell(row=row, column=4).value
    sd_name = ws.cell(row=row, column=5).value
    ctrl_name = ws.cell(row=row, column=7).value
    criteria_design = ws.cell(row=row, column=8).value
    criteria_effect = ws.cell(row=row, column=9).value

    if not domain_code or not ctrl_name:
        continue

    domain_code = str(domain_code).strip()
    ctrl_name_str = str(ctrl_name).strip()
    sd_code_str = str(sd_code or '').strip()

    # Get control number from the map sheet
    ctrl_num = ctrl_num_map.get(ctrl_name_str)
    if not ctrl_num:
        continue  # skip if no control number found

    # Domains
    if domain_code not in domains:
        domains[domain_code] = str(domain_name or '').strip()

    # Subdomains
    if sd_code_str and sd_code_str not in subdomains:
        subdomains[sd_code_str] = {
            'name': str(sd_name or '').strip(),
            'domain_code': domain_code,
        }

    # Controls
    if ctrl_num not in controls:
        controls[ctrl_num] = {
            'num': ctrl_num,
            'name': ctrl_name_str,
            'sd_code': sd_code_str,
            'domain_code': domain_code,
            'design': str(criteria_design or '').strip(),
            'effect': str(criteria_effect or '').strip(),
        }


def esc(s):
    """Escape single quotes for SQL."""
    return s.replace("'", "''")


sql_lines = []

# Domain nodes (depth=0)
domain_ids = {}
domain_sort = 0
for code, name in domains.items():
    domain_sort += 1
    did = str(uuid.uuid4())
    domain_ids[code] = did
    sql_lines.append(
        f"INSERT INTO framework_nodes "
        f"(id, framework_id, parent_id, name, reference_code, node_type, "
        f"is_assessable, is_active, sort_order, path, depth, created_at, updated_at) "
        f"VALUES ('{did}', '{FW_ID}', NULL, '{esc(name)}', "
        f"'{code}', 'domain', false, true, {domain_sort}, "
        f"'{code}', 0, NOW(), NOW());"
    )

# Subdomain nodes (depth=1)
subdomain_ids = {}
sd_sort = 0
for code, sd in subdomains.items():
    sd_sort += 1
    sid = str(uuid.uuid4())
    subdomain_ids[code] = sid
    parent_id = domain_ids[sd['domain_code']]
    path = f"{sd['domain_code']}/{code}"
    sql_lines.append(
        f"INSERT INTO framework_nodes "
        f"(id, framework_id, parent_id, name, reference_code, node_type, "
        f"is_assessable, is_active, sort_order, path, depth, created_at, updated_at) "
        f"VALUES ('{sid}', '{FW_ID}', '{parent_id}', '{esc(sd['name'])}', "
        f"'{code}', 'subdomain', false, true, {sd_sort}, "
        f"'{path}', 1, NOW(), NOW());"
    )

# Control nodes (depth=2)
ctrl_sort = 0
for ctrl_num, c in controls.items():
    ctrl_sort += 1
    cid = str(uuid.uuid4())
    parent_id = subdomain_ids.get(c['sd_code'])
    if not parent_id:
        continue  # skip controls without a valid subdomain

    path = f"{c['domain_code']}/{c['sd_code']}/{ctrl_num}"

    # Build acceptance_criteria from design + effectiveness criteria
    criteria_parts = []
    if c['design']:
        criteria_parts.append(f"Test of Design:\n{c['design']}")
    if c['effect']:
        criteria_parts.append(f"Test of Effectiveness:\n{c['effect']}")
    acceptance_criteria = '\n\n'.join(criteria_parts)

    sql_lines.append(
        f"INSERT INTO framework_nodes "
        f"(id, framework_id, parent_id, name, reference_code, node_type, "
        f"is_assessable, is_active, sort_order, path, depth, "
        f"acceptance_criteria, created_at, updated_at) "
        f"VALUES ('{cid}', '{FW_ID}', '{parent_id}', '{esc(c['name'])}', "
        f"'{esc(ctrl_num)}', 'control', true, true, {ctrl_sort}, "
        f"'{esc(path)}', 2, "
        f"'{esc(acceptance_criteria)}', NOW(), NOW());"
    )

# Write SQL file
with open(SQL_OUTPUT, 'w', encoding='utf-8') as f:
    f.write('BEGIN;\n')
    for line in sql_lines:
        f.write(line + '\n')
    f.write('COMMIT;\n')

print(f"Generated {len(sql_lines)} SQL inserts:")
print(f"  {len(domains)} domains")
print(f"  {len(subdomains)} subdomains")
ctrl_count = len(sql_lines) - len(domains) - len(subdomains)
print(f"  {ctrl_count} controls")
print(f"SQL written to {SQL_OUTPUT}")
