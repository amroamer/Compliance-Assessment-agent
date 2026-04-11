#!/usr/bin/env python3
"""Load AI Badges framework hierarchy from Excel into the database."""
import sys
import io
import uuid

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

EXCEL_PATH = r'C:\Users\amroa\Downloads\AI_Ethics_Documentation_Template.xlsx'
FW_ID = '33f661cb-69a0-4c9e-8c3c-94ab4ce40e17'
SQL_OUTPUT = r'C:\projects\AI-Badges\scripts\ai_badges_insert.sql'

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['Requirements Summary']

# Parse the sheet - Report and Requirements use merged cells
current_report = None
current_req = None
reports = {}  # ordered dict of report_name -> {reqs: ordered dict of req_name -> [outcomes]}
report_order = []

for row in range(2, ws.max_row + 1):
    report = ws.cell(row=row, column=1).value
    req = ws.cell(row=row, column=2).value
    outcome = ws.cell(row=row, column=3).value

    if report:
        current_report = str(report).strip()
        if current_report not in reports:
            reports[current_report] = {}
            report_order.append(current_report)
    if req:
        current_req = str(req).strip()
        if current_report and current_req not in reports[current_report]:
            reports[current_report][current_req] = []
    if outcome and current_report and current_req:
        reports[current_report][current_req].append(str(outcome).strip())


def esc(s):
    """Escape single quotes for SQL."""
    return s.replace("'", "''")


sql_lines = []
report_sort = 0
req_global_sort = 0

for rpt_name in report_order:
    reqs = reports[rpt_name]
    report_sort += 1
    rid = str(uuid.uuid4())

    # Build a short reference code for the report (R1, R2, ...)
    ref = f'R{report_sort}'
    path = ref

    sql_lines.append(
        f"INSERT INTO framework_nodes "
        f"(id, framework_id, parent_id, name, reference_code, node_type, "
        f"is_assessable, is_active, sort_order, path, depth, created_at, updated_at) "
        f"VALUES ('{rid}', '{FW_ID}', NULL, '{esc(rpt_name)}', "
        f"'{ref}', 'Report', false, true, {report_sort}, "
        f"'{path}', 0, NOW(), NOW());"
    )

    req_sort = 0
    for req_name, outcomes in reqs.items():
        req_sort += 1
        req_global_sort += 1
        qid = str(uuid.uuid4())

        # Reference code: R1.1, R1.2, etc.
        req_ref = f'{ref}.{req_sort}'
        req_path = f'{path}/{req_ref}'

        # acceptance_criteria = concatenated expected outcomes
        acceptance_criteria = '\n'.join(f'- {o}' for o in outcomes)

        sql_lines.append(
            f"INSERT INTO framework_nodes "
            f"(id, framework_id, parent_id, name, reference_code, node_type, "
            f"is_assessable, is_active, sort_order, path, depth, "
            f"acceptance_criteria, created_at, updated_at) "
            f"VALUES ('{qid}', '{FW_ID}', '{rid}', '{esc(req_name)}', "
            f"'{req_ref}', 'Requirement', true, true, {req_sort}, "
            f"'{req_path}', 1, "
            f"'{esc(acceptance_criteria)}', NOW(), NOW());"
        )

# Write SQL file
with open(SQL_OUTPUT, 'w', encoding='utf-8') as f:
    f.write('BEGIN;\n')
    for line in sql_lines:
        f.write(line + '\n')
    f.write('COMMIT;\n')

total_reqs = sum(len(reqs) for reqs in reports.values())
print(f"Generated {len(sql_lines)} SQL inserts:")
print(f"  {len(reports)} reports")
print(f"  {total_reqs} requirements")
print(f"SQL written to {SQL_OUTPUT}")
